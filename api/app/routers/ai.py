"""
Routeur SahamAI : chatbot analytique en langage naturel.

POST /ai/ask   -> question en français -> réponse + SQL + tableau + graphe,
                  journalisée dans ai_query_log (page Admin).
GET  /ai/logs  -> historique des questions posées (pour la console admin).
POST /ai/export/pdf -> génération de rapport PDF professionnel
POST /ai/export/excel -> export des données en Excel

Protégé par JWT : il faut être connecté (n'importe quel rôle peut poser
une question, mais seul ADMIN/DG voit le log complet).
"""

from typing import Any, Dict, List, Optional
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, field_validator
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.deps import get_current_user
from app.database import engine, get_db
from app.models.user import User
from app.services.rag.text2sql import answer_question

router = APIRouter(prefix="/ai", tags=["SahamAI"])


class AskRequest(BaseModel):
    question: str

    @field_validator("question")
    @classmethod
    def question_not_blank(cls, v: str) -> str:
        v = v.strip()
        if not v:
            raise ValueError("La question ne peut pas être vide")
        return v


class BoxPlotItem(BaseModel):
    label: str
    min: float
    q1: float
    median: float
    q3: float
    max: float
    outliers: Optional[List[float]] = None


class ChartData(BaseModel):
    type: str
    title: Optional[str] = None
    labels: Optional[List[str]] = None
    values: Optional[List[float]] = None
    unit: Optional[str] = None
    series: Optional[List[Dict[str, Any]]] = None
    boxes: Optional[List[BoxPlotItem]] = None
    tool_used: Optional[str] = None


class AskResponse(BaseModel):
    mode: str
    answer: str
    sql: Optional[str] = None
    sql_explanation: Optional[str] = None
    columns: Optional[List[str]] = None
    rows: Optional[List[List[Any]]] = None
    row_count: Optional[int] = None
    chart: Optional[ChartData] = None
    tables: Optional[List[str]] = None
    duration_ms: Optional[int] = None
    # Vrai quand la reponse vient du mode de secours (LLM injoignable).
    # L'interface l'affiche honnetement au lieu de faire passer une reponse
    # generique pour une vraie analyse.
    degraded: Optional[bool] = False
    from_cache: Optional[bool] = False


class ExportRequest(BaseModel):
    chat_history: List[Dict[str, Any]]
    title: str = "Rapport Analytique Saham Bank"
    include_charts: bool = True
    include_sql: bool = False
    author: str = ""


@router.post("/ask", response_model=AskResponse)
async def ask(
    body: AskRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return answer_question(db, user, body.question)


@router.get("/logs")
async def logs(
    limit: int = 50,
    user: User = Depends(get_current_user),
):
    limit = max(1, min(limit, 200))
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT id, user_nom, user_id, question, mode, sql_generated,
                       row_count, duration_ms, status, error, answer, tables,
                       created_at
                FROM ai_query_log
                ORDER BY created_at DESC, id DESC
                LIMIT :limit
                """
            ),
            {"limit": limit},
        ).fetchall()
    return [
        {
            "id": r[0], "user_nom": r[1], "user_id": r[2], "question": r[3],
            "mode": r[4], "sql": r[5], "row_count": r[6], "duration_ms": r[7],
            "status": r[8], "error": r[9], "answer": r[10], "tables": r[11],
            "created_at": r[12].isoformat() if r[12] else None,
        }
        for r in rows
    ]


@router.post("/export/pdf")
async def export_pdf(
    body: ExportRequest,
    user: User = Depends(get_current_user),
):
    """Génère un rapport PDF professionnel à partir de l'historique de chat."""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        # Création du buffer PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a365d'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2c5282'),
            spaceAfter=12,
            spaceBefore=20
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=12,
            leading=14
        )
        
        # Contenu du document
        story = []
        
        # En-tête
        story.append(Paragraph("SAHAM BANK", ParagraphStyle('BankName', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor('#1a365d'), alignment=TA_CENTER)))
        story.append(Spacer(1, 0.2*inch))
        story.append(Paragraph(body.title, title_style))
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=9, textColor=colors.gray, alignment=TA_CENTER)))
        if body.author:
            story.append(Paragraph(f"Par {body.author}", ParagraphStyle('AuthorStyle', parent=styles['Normal'], fontSize=9, textColor=colors.gray, alignment=TA_CENTER)))
        story.append(Spacer(1, 0.3*inch))
        
        # Contenu du chat
        for i, msg in enumerate(body.chat_history, 1):
            if msg.get('role') == 'user':
                story.append(Paragraph(f"Question {i} : {msg.get('content', '')}", heading_style))
            elif msg.get('role') == 'assistant':
                answer = msg.get('content', '')
                if answer:
                    story.append(Paragraph(answer, normal_style))
                
                # Ajouter les données si disponibles
                if 'data' in msg and msg['data']:
                    data = msg['data']
                    
                    # Tableau de données
                    if 'columns' in data and 'rows' in data and data['rows']:
                        table_data = [data['columns']] + data['rows']
                        table = Table(table_data, repeatRows=1)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e2e8f0')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        story.append(table)
                        story.append(Spacer(1, 0.2*inch))
                    
                    # SQL si demandé
                    if body.include_sql and 'sql' in data and data['sql']:
                        story.append(Paragraph("Requête SQL :", ParagraphStyle('SQLLabel', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', textColor=colors.HexColor('#718096'))))
                        story.append(Paragraph(data['sql'], ParagraphStyle('SQLCode', parent=styles['Normal'], fontSize=8, fontName='Courier', textColor=colors.HexColor('#2d3748'), backgroundColor=colors.HexColor('#f7fafc'))))
                        story.append(Spacer(1, 0.2*inch))
                
                story.append(Spacer(1, 0.2*inch))
        
        # Pied de page
        story.append(PageBreak())
        story.append(Paragraph("Document généré par SahamAI - Assistant Analytique Saham Bank", ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray, alignment=TA_CENTER)))
        
        # Génération du PDF
        doc.build(story)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=rapport_saham_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Module reportlab non installé. Installez-le avec: pip install reportlab")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération PDF: {str(e)}")


@router.post("/export/excel")
async def export_excel(
    body: ExportRequest,
    user: User = Depends(get_current_user),
):
    """Exporte les données du chat en fichier Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Création du workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport Analytique"
        
        # En-tête
        ws['A1'] = "SAHAM BANK - Rapport Analytique"
        ws['A1'].font = Font(size=16, bold=True, color="1a365d")
        ws['A2'] = f"Titre: {body.title}"
        ws['A2'].font = Font(size=12, bold=True)
        ws['A3'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        if body.author:
            ws['A4'] = f"Auteur: {body.author}"
        
        row_num = 6
        
        # Contenu du chat
        for i, msg in enumerate(body.chat_history, 1):
            if msg.get('role') == 'user':
                ws[f'A{row_num}'] = f"Question {i}: {msg.get('content', '')}"
                ws[f'A{row_num}'].font = Font(bold=True, size=11, color="2c5282")
                row_num += 1
                
            elif msg.get('role') == 'assistant':
                answer = msg.get('content', '')
                if answer:
                    ws[f'A{row_num}'] = f"Réponse: {answer}"
                    ws[f'A{row_num}'].font = Font(size=10)
                    row_num += 1
                
                # Données si disponibles
                if 'data' in msg and msg['data']:
                    data = msg['data']
                    
                    if 'columns' in data and 'rows' in data and data['rows']:
                        row_num += 1
                        ws[f'A{row_num}'] = "Données:"
                        ws[f'A{row_num}'].font = Font(bold=True)
                        row_num += 1
                        
                        # En-têtes de tableau
                        for col_idx, col_name in enumerate(data['columns'], 1):
                            cell = ws.cell(row=row_num, column=col_idx, value=col_name)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="4a5568", end_color="4a5568", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center")
                        
                        row_num += 1
                        
                        # Données
                        for row_data in data['rows']:
                            for col_idx, value in enumerate(row_data, 1):
                                ws.cell(row=row_num, column=col_idx, value=value)
                            row_num += 1
                        
                        # Ajustement des colonnes
                        for col_idx in range(1, len(data['columns']) + 1):
                            ws.column_dimensions[get_column_letter(col_idx)].auto_size = True
                    
                    # SQL si demandé
                    if body.include_sql and 'sql' in data and data['sql']:
                        row_num += 1
                        ws[f'A{row_num}'] = "SQL:"
                        ws[f'A{row_num}'].font = Font(bold=True, color="718096")
                        row_num += 1
                        ws[f'A{row_num}'] = data['sql']
                        ws[f'A{row_num}'].font = Font(name="Courier", size=9, color="2d3748")
                        row_num += 1
                
                row_num += 2
        
        # Sauvegarde dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=rapport_saham_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Module openpyxl non installé. Installez-le avec: pip install openpyxl")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'export Excel: {str(e)}")
